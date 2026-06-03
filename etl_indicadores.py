"""
ETL Indicadores de Trayectoria Escolar
"""
from __future__ import annotations
import logging, os, re, sys
from pathlib import Path
import pyodbc
from dotenv import load_dotenv
from openpyxl import load_workbook

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout), logging.FileHandler("etl_indicadores.log", encoding="utf-8")],
)
logger = logging.getLogger("etl.indicadores")

MAPA_INDICADOR = [(r"repitencia","REP"),(r"abandono","ABN"),(r"sobreedad","SOB")]
MAPA_NIVEL = {"primaria":"PRI","primario":"PRI","secundaria":"SEC","secundario":"SEC"}
PATRON_RUIDO = re.compile(
    r"^(total\s+pa[ií]s|tasa\s+de|fuente|realizaci[oó]n|ministerio|secretar|direcci[oó]n|"
    r"red\s+federal|subsecretar|nota|pol[ií]tico|division|ley\s+n|estructura\s+educativa|"
    r"educaci[oó]n\s+primaria\s+de|educaci[oó]n\s+secundaria\s+de)", re.IGNORECASE)

MERGE_SQL = """
MERGE dbo.hechos_indicadores_trayectoria AS d
USING (SELECT ? AS indicador_id,? AS anio_lectivo,? AS jurisdiccion_id,
              ? AS nivel_id,? AS anio_estudio,? AS valor,? AS fuente_archivo) AS o
ON (d.indicador_id=o.indicador_id AND d.anio_lectivo=o.anio_lectivo AND
    d.jurisdiccion_id=o.jurisdiccion_id AND d.nivel_id=o.nivel_id AND d.anio_estudio=o.anio_estudio)
WHEN MATCHED THEN UPDATE SET d.valor=o.valor,d.fuente_archivo=o.fuente_archivo,d.fecha_ingesta=SYSUTCDATETIME()
WHEN NOT MATCHED THEN INSERT (indicador_id,anio_lectivo,jurisdiccion_id,nivel_id,anio_estudio,valor,fuente_archivo)
VALUES (o.indicador_id,o.anio_lectivo,o.jurisdiccion_id,o.nivel_id,o.anio_estudio,o.valor,o.fuente_archivo);"""

def get_conn():
    load_dotenv()
    server=os.getenv("SQL_SERVER","localhost"); database=os.getenv("SQL_DATABASE","educacion_argentina")
    user=os.getenv("SQL_USER",""); password=os.getenv("SQL_PASSWORD","")
    drivers=[d for d in pyodbc.drivers() if "SQL Server" in d]
    driver=sorted(drivers,key=lambda d:int(re.search(r"\d+",d).group()) if re.search(r"\d+",d) else 0,reverse=True)[0]
    cs=f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};UID={user};PWD={password};" if user else f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};Trusted_Connection=yes;"
    conn=pyodbc.connect(cs); conn.autocommit=False; return conn

def cargar_caches(conn):
    c=conn.cursor()
    c.execute("SELECT codigo,indicador_id FROM dbo.dim_indicadores")
    cache_ind={r[0]:r[1] for r in c.fetchall()}
    c.execute("SELECT codigo_redfie,nivel_id FROM dbo.dim_niveles_educativos")
    cache_niv={r[0]:r[1] for r in c.fetchall()}
    c.execute("SELECT LOWER(RTRIM(LTRIM(nombre_oficial))),jurisdiccion_id FROM dbo.dim_jurisdicciones UNION SELECT LOWER(RTRIM(LTRIM(nombre_corto))),jurisdiccion_id FROM dbo.dim_jurisdicciones WHERE nombre_corto IS NOT NULL")
    cache_jur={r[0]:r[1] for r in c.fetchall()}
    alias={"resto de bs as":cache_jur.get("buenos aires resto"),"resto de los partidos":cache_jur.get("buenos aires resto"),
           "conurbano":cache_jur.get("gba - conurbano"),"cordoba":cache_jur.get("córdoba"),
           "entre rios":cache_jur.get("entre ríos"),"neuquen":cache_jur.get("neuquén"),
           "rio negro":cache_jur.get("río negro"),"tucuman":cache_jur.get("tucumán"),
           "tierra del fuego":cache_jur.get("tierra del fuego")}
    for k,v in alias.items():
        if v and k not in cache_jur: cache_jur[k]=v
    return cache_ind,cache_niv,cache_jur

def detectar_indicador(nombre):
    for patron,codigo in MAPA_INDICADOR:
        if re.search(patron,nombre.lower()): return codigo
    return None

def detectar_anio(hoja):
    m=re.match(r"(\d{4})-(\d{2,4})$",hoja.strip())
    if m: return int(m.group(1))
    m=re.match(r"(\d{4})$",hoja.strip())
    if m: return int(m.group(1))
    return None

def extraer_hoja(ws, anio, codigo_indicador, nombre_archivo):
    """
    Formato division territorial 2012-2025.
    Col 0 = jurisdiccion, Col 1 = estructura educativa (6-6, etc.), Col 2+ = datos.
    Fila de header tiene Primaria / Secundaria como labels de grupos de columnas.
    """
    filas=list(ws.iter_rows(values_only=True))
    resultados=[]
    PATRON_NIVEL=re.compile(r"(primaria|primario|secundaria|secundario)",re.IGNORECASE)

    # Buscar fila donde col 0 tiene "Division" y otras cols tienen "Primaria"/"Secundaria"
    fila_niveles_idx=None; fila_anios_idx=None
    for i,row in enumerate(filas[:20]):
        if row[0] and re.search(r"divis",str(row[0]),re.IGNORECASE):
            if any(v and PATRON_NIVEL.search(str(v)) for v in row):
                fila_niveles_idx=i
        row_text=" ".join(str(v) for v in row if v)
        if fila_anios_idx is None and re.search(r"1[°º]",row_text,re.IGNORECASE):
            fila_anios_idx=i
        if fila_anios_idx is None and re.search(r"total.*1[°º]",row_text,re.IGNORECASE):
            fila_anios_idx=i

    # Caso especial: header en 3 filas (2020/2021 sobreedad)
    # fila_niveles tiene Primaria/Secundaria pero fila_anios no tiene cols numericas de datos
    if fila_niveles_idx is not None and fila_anios_idx is not None:
        fila_an_test = filas[fila_anios_idx]
        # Si la fila de anios no tiene Total en col 2, los datos estan en la fila siguiente
        if fila_anios_idx + 1 < len(filas):
            next_row = filas[fila_anios_idx + 1]
            next_text = " ".join(str(v) for v in next_row if v)
            if re.search(r"1[°º]", next_text, re.IGNORECASE) and not re.search(r"1[°º]", " ".join(str(v) for v in fila_an_test if v), re.IGNORECASE):
                fila_anios_idx = fila_anios_idx + 1

    col_nivel_map={}
    if fila_niveles_idx is not None and fila_anios_idx is not None:
        fila_niv=filas[fila_niveles_idx]; fila_an=filas[fila_anios_idx]
        # Si fila_anios no tiene Total (caso 2020/2021), buscar Total en fila intermedia
        fila_total_extra = None
        if fila_anios_idx > fila_niveles_idx + 1:
            for fi in range(fila_niveles_idx+1, fila_anios_idx):
                row_text = " ".join(str(v) for v in filas[fi] if v)
                if re.search(r"total", row_text, re.IGNORECASE):
                    fila_total_extra = filas[fi]
                    break
        nivel_actual=None
        for col_idx in range(max(len(fila_niv), len(fila_an))):
            val_niv=fila_niv[col_idx] if col_idx<len(fila_niv) else None
            if val_niv and PATRON_NIVEL.search(str(val_niv)):
                m=PATRON_NIVEL.search(str(val_niv)); nivel_actual=MAPA_NIVEL.get(m.group(1).lower())
            val_an=fila_an[col_idx] if col_idx<len(fila_an) else None
            # Si no hay valor en fila_an, buscar en fila_total_extra
            if not val_an and fila_total_extra and col_idx<len(fila_total_extra):
                val_an=fila_total_extra[col_idx]
            if val_an and nivel_actual:
                an_str=str(val_an).strip()
                if re.search(r"(total|[1-9][°º]|[1-9]\s*año)",an_str,re.IGNORECASE):
                    if not re.search(r"año\s+de\s+estudio|estructura",an_str,re.IGNORECASE):
                        col_nivel_map[col_idx]=(nivel_actual,an_str[:30])
    elif fila_anios_idx is not None:
        fila_an=filas[fila_anios_idx]; codigo_nivel_unico=None
        for i2 in range(max(0,fila_anios_idx-5),fila_anios_idx):
            row_text=" ".join(str(v) for v in filas[i2] if v)
            m=PATRON_NIVEL.search(row_text)
            if m: codigo_nivel_unico=MAPA_NIVEL.get(m.group(1).lower()); break
        if codigo_nivel_unico:
            for col_idx,val in enumerate(fila_an):
                if val:
                    an_str=str(val).strip()
                    if re.search(r"(total|[1-9][°º]|[1-9]\s*año)",an_str,re.IGNORECASE):
                        if not re.search(r"año\s+de\s+estudio|estructura",an_str,re.IGNORECASE):
                            col_nivel_map[col_idx]=(codigo_nivel_unico,an_str[:30])

    if not col_nivel_map: return resultados
    datos_inicio=(fila_anios_idx or 0)+1

    for fila in filas[datos_inicio:]:
        jur_raw=str(fila[0]).strip() if fila[0] else ""
        if not jur_raw or PATRON_RUIDO.match(jur_raw): continue
        jur_clean=re.sub(r"\s+"," ",jur_raw).strip()
        for col_idx,(codigo_nivel,anio_label) in col_nivel_map.items():
            if col_idx>=len(fila): continue
            val=fila[col_idx]
            if val is None: continue
            try:
                valor=float(val)
                resultados.append({"codigo_indicador":codigo_indicador,"anio_lectivo":anio,
                    "jurisdiccion_raw":jur_clean,"codigo_nivel":codigo_nivel,
                    "anio_estudio":anio_label,"valor":round(valor,4),"fuente_archivo":nombre_archivo})
            except (TypeError,ValueError): pass
    return resultados

def extraer_hoja_homogenea(ws, anio, codigo_indicador, nombre_archivo):
    """
    Formato homogeneo 2003-2016.
    Estructura: col 0 = jurisdiccion, col 1 = Total nivel, col 2+ = años de estudio.
    Header: fila 10 = nivel, fila 11 = 'Total'+'Año de estudio', fila 12 = '1°','2°'...
    """
    filas=list(ws.iter_rows(values_only=True))
    resultados=[]
    PATRON_NIVEL=re.compile(r"(primaria|primario|secundaria|secundario)",re.IGNORECASE)
    PATRON_JUR=re.compile(
        r"^(tasa\s+de|fuente|realizaci[oó]n|ministerio|direcci[oó]n|red\s+federal|"
        r"educacion\s+comun|nota|secretar|subsecretar|div[^\s])",re.IGNORECASE)

    codigo_nivel=None; col_total=None; col_anios=[]; datos_inicio=None

    for i,row in enumerate(filas):
        row_text=" ".join(str(v) for v in row if v)

        # Detectar nivel desde col 1 o col 2
        for col_niv in [1, 2]:
            if len(row)>col_niv and row[col_niv]:
                m=PATRON_NIVEL.search(str(row[col_niv]))
                if m: codigo_nivel=MAPA_NIVEL.get(m.group(1).lower()); break

        # Detectar fila de anos de estudio (tiene '1°' o '1º')
        if re.search(r"1[°º]|1º",row_text,re.IGNORECASE) and codigo_nivel:
            # Los años están en fila actual o fila siguiente
            # Buscar columna de Total en fila actual o anterior
            fila_total = filas[i-1] if i > 0 else row
            for col_idx,val in enumerate(fila_total):
                if val and re.search(r"^total$",str(val).strip(),re.IGNORECASE):
                    col_total=col_idx; break
            if col_total is None:
                col_total=1  # default: Total siempre en col 1

            # Años de estudio en fila actual
            for col_idx,val in enumerate(row):
                if val and re.search(r"[1-9][°º]|[1-9]º",str(val),re.IGNORECASE):
                    col_anios.append((col_idx,str(val).strip()[:30]))

            datos_inicio=i+1
            break

    if not codigo_nivel or datos_inicio is None:
        return []

    if col_total is None:
        col_total=1

    for fila in filas[datos_inicio:]:
        # Jurisdiccion en col 0 o col 1 (algunos años tienen col 0 vacia)
        jur_raw = ""
        if fila[0]:
            jur_raw = str(fila[0]).strip()
        elif len(fila) > 1 and fila[1]:
            jur_raw = str(fila[1]).strip()
            # Si estamos leyendo de col 1, ajustar col_total y col_anios
            # (esto ocurre en hojas 2007-2009 del formato viejo)
        if not jur_raw or PATRON_JUR.match(jur_raw):
            continue
        if len(jur_raw) < 3:
            continue
        jur_clean=re.sub(r"\s+"," ",jur_raw).strip()

        # Total del nivel
        if col_total < len(fila) and fila[col_total] is not None:
            try:
                valor=float(fila[col_total])
                resultados.append({"codigo_indicador":codigo_indicador,"anio_lectivo":anio,
                    "jurisdiccion_raw":jur_clean,"codigo_nivel":codigo_nivel,
                    "anio_estudio":"Total","valor":round(valor,4),"fuente_archivo":nombre_archivo})
            except (TypeError,ValueError): pass

        # Años de estudio
        for col_idx,label in col_anios:
            if col_idx<len(fila) and fila[col_idx] is not None:
                try:
                    valor=float(fila[col_idx])
                    resultados.append({"codigo_indicador":codigo_indicador,"anio_lectivo":anio,
                        "jurisdiccion_raw":jur_clean,"codigo_nivel":codigo_nivel,
                        "anio_estudio":label,"valor":round(valor,4),"fuente_archivo":nombre_archivo})
                except (TypeError,ValueError): pass

    return resultados

def cargar_registros(registros, conn, caches):
    cache_ind,cache_niv,cache_jur=caches
    cursor=conn.cursor(); insertadas=0; errores=0
    for r in registros:
        ind_id=cache_ind.get(r["codigo_indicador"])
        niv_id=cache_niv.get(r["codigo_nivel"])
        jur_id=cache_jur.get(r["jurisdiccion_raw"].lower())
        if not ind_id or not niv_id or not jur_id:
            if not jur_id: logger.debug("Jur no mapeada: '%s'",r["jurisdiccion_raw"])
            errores+=1; continue
        try:
            cursor.execute(MERGE_SQL,ind_id,int(r["anio_lectivo"]),jur_id,niv_id,
                           r["anio_estudio"],r["valor"],r["fuente_archivo"])
            insertadas+=1
        except pyodbc.Error as e:
            logger.error("SQL error: %s",e); errores+=1
    return insertadas,errores

def procesar_archivo(path, conn, caches):
    codigo_indicador=detectar_indicador(path.name)
    if not codigo_indicador:
        logger.warning("No se pudo detectar indicador en: %s",path.name); return 0,0
    logger.info("Procesando: %s (%s)",path.name,codigo_indicador)
    es_homogeneo="homogenea" in path.name.lower() or "homog" in path.name.lower()
    wb=load_workbook(path,read_only=True,data_only=True)
    total_ins=0; total_err=0
    for nombre_hoja in wb.sheetnames:
        if nombre_hoja.lower() in ("limitaciones","metodologia","notas"): continue
        anio=detectar_anio(nombre_hoja)
        if not anio: continue
        ws=wb[nombre_hoja]
        registros=extraer_hoja_homogenea(ws,anio,codigo_indicador,path.name) if es_homogeneo else extraer_hoja(ws,anio,codigo_indicador,path.name)
        if not registros:
            logger.warning("  Hoja %s: sin datos",nombre_hoja); continue
        ins,err=cargar_registros(registros,conn,caches)
        total_ins+=ins; total_err+=err
        logger.info("  Hoja %s (año %d): %d registros, %d errores",nombre_hoja,anio,ins,err)
    return total_ins,total_err

def procesar_directorio(directorio):
    base=Path(directorio)
    if not base.exists():
        logger.error("Directorio no existe: %s",directorio); sys.exit(1)
    archivos=sorted(base.glob("*.xlsx"))+sorted(base.glob("*.XLSX"))
    if not archivos:
        logger.error("No se encontraron archivos XLSX en %s",directorio); sys.exit(1)
    logger.info("="*65)
    logger.info("ETL INDICADORES TRAYECTORIA — %s",directorio)
    logger.info("Archivos encontrados: %d",len(archivos))
    logger.info("="*65)
    conn=get_conn(); caches=cargar_caches(conn)
    total_global=0
    for archivo in archivos:
        ins,err=procesar_archivo(archivo,conn,caches)
        conn.commit(); total_global+=ins
        logger.info("  -> %d insertados, %d errores\n",ins,err)
    conn.close()
    logger.info("="*65)
    logger.info("TOTAL INSERTADOS: %d",total_global)
    logger.info("="*65)

if __name__=="__main__":
    if len(sys.argv)!=2:
        print("Uso: python etl_indicadores.py <carpeta_xlsx>"); sys.exit(1)
    load_dotenv()
    procesar_directorio(sys.argv[1])
