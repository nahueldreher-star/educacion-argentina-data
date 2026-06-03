"""
ETL Histórico — Sistema Educativo Argentino
Procesa anuarios de 2003 a 2024 desde una carpeta local.

Soporta:
- ZIP con XLSX (2015-2024)
- RAR con XLS (2003-2014)
- Conversión automática XLS → XLSX via LibreOffice o COM

Uso:
    python etl_historico.py <carpeta_anuarios>
    python etl_historico.py "C:\\Users\\Administrator\\Desktop\\Anuarios"
"""

from __future__ import annotations

import logging
import os
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

import pandas as pd
import pyodbc
from dotenv import load_dotenv

# ==============================================================================
# LOGGING
# ==============================================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("etl_historico.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger("etl.historico")

# ==============================================================================
# CONFIGURACION
# ==============================================================================

SIETE_ZIP = r"C:\Program Files\7-Zip\7z.exe"

# Patrones para detectar el nivel desde el nombre del archivo
MAPA_NIVEL = [
    (r"inicial|jardin|preescolar",       "Educacion Inicial",                 "INI"),
    (r"primar",                           "Educacion Primaria",                "PRI"),
    (r"secundar|medio|polimodal",         "Educacion Secundaria",              "SEC"),
    (r"superior",                         "Educacion Superior No Universitaria","SUP"),
    (r"especial",                         "Educacion Especial",                "ESP"),
    (r"adultos|epja",                     "Educacion Permanente EPJA",         "PER"),
]

# Nombres de hoja de alumnos según año
# Antes de 2015: 'alumnos' (minúscula)
# 2015-2023: 'Alu_Total' o 'Alumnos'
# 2024-2025: 'Alu_Total' o 'Alumnos'
HOJAS_CANDIDATAS = ["Alu_Total", "Alumnos", "alumnos", "ALUMNOS", "Alu_total"]

# Hojas de matricula por año (formato 2016+)
HOJAS_MATRICULA_PATRON = re.compile(
    r"(matrícula|matricula|alumnos?)\s*(est\.?\s*p|total|por)?",
    re.IGNORECASE
)

PATRON_ESTATAL = re.compile(
    r"sector\s+de\s+gesti[oó]n\s+estatal|sector\s+estatal|gesti[oó]n\s+estatal",
    re.IGNORECASE,
)
PATRON_PRIVADO = re.compile(
    r"sector\s+de\s+gesti[oó]n\s+privad|sector\s+privado|gesti[oó]n\s+privad",
    re.IGNORECASE,
)
PATRON_HEADER = re.compile(r"^divisi[oó]n\s+pol[ií]tico", re.IGNORECASE)
PATRON_RUIDO = re.compile(
    r"^(total\s+pa[ií]s|total\s+general|nota:|fuente:|realizaci[oó]n:|incluye|"
    r"a[ñn]o\s+\d{4}|la\s+resoluci[oó]n|el\s+tipo\s+de|ministerio|direcci[oó]n|"
    r"red\s+federal|realización)",
    re.IGNORECASE,
)

ETL_VERSION = "2.0-historico"

MERGE_SQL = """
MERGE dbo.hechos_matricula_historica AS destino
USING (
    SELECT ? AS anio_lectivo, ? AS jurisdiccion_id, ? AS nivel_id,
           ? AS sector_id, ? AS total_alumnos, ? AS fuente_archivo, ? AS version_etl
) AS origen
ON (destino.anio_lectivo = origen.anio_lectivo AND
    destino.jurisdiccion_id = origen.jurisdiccion_id AND
    destino.nivel_id = origen.nivel_id AND
    destino.sector_id = origen.sector_id)
WHEN MATCHED THEN
    UPDATE SET destino.total_alumnos = origen.total_alumnos,
               destino.fuente_archivo = origen.fuente_archivo,
               destino.fecha_ingesta = SYSUTCDATETIME(),
               destino.version_etl = origen.version_etl
WHEN NOT MATCHED THEN
    INSERT (anio_lectivo, jurisdiccion_id, nivel_id, sector_id,
            total_alumnos, fuente_archivo, version_etl)
    VALUES (origen.anio_lectivo, origen.jurisdiccion_id, origen.nivel_id,
            origen.sector_id, origen.total_alumnos, origen.fuente_archivo,
            origen.version_etl);
"""

# ==============================================================================
# UTILIDADES
# ==============================================================================

def inferir_anio(ruta: str) -> int | None:
    """Infiere el año lectivo desde la ruta del archivo o carpeta padre."""
    texto = Path(ruta).stem + " " + str(Path(ruta).parent)
    match = re.search(r"(19[9]\d|20[0-2]\d)", texto.lower())
    return int(match.group(1)) if match else None


def inferir_nivel(nombre: str) -> tuple[str, str] | None:
    """Infiere nivel educativo desde el nombre del archivo."""
    nombre_lower = nombre.lower()
    for patron, nombre_nivel, codigo in MAPA_NIVEL:
        if re.search(patron, nombre_lower):
            return nombre_nivel, codigo
    return None


def descomprimir(archivo: Path, destino: Path) -> bool:
    """Descomprime ZIP o RAR usando 7-Zip."""
    if not Path(SIETE_ZIP).exists():
        logger.error("7-Zip no encontrado en %s", SIETE_ZIP)
        return False
    try:
        result = subprocess.run(
            [SIETE_ZIP, "x", str(archivo), f"-o{destino}", "-y"],
            capture_output=True, text=True, timeout=120
        )
        if result.returncode == 0:
            logger.info("Descomprimido: %s", archivo.name)
            return True
        else:
            logger.error("Error descomprimiendo %s: %s", archivo.name, result.stderr[:200])
            return False
    except subprocess.TimeoutExpired:
        logger.error("Timeout descomprimiendo %s", archivo.name)
        return False


def convertir_xls_a_xlsx(xls_path: Path, destino_dir: Path) -> Path | None:
    """Convierte XLS a XLSX usando win32com (Excel COM) o LibreOffice."""
    xlsx_path = destino_dir / (xls_path.stem + ".xlsx")

    # Intentar con win32com (Excel nativo en Windows)
    try:
        import win32com.client
        import pythoncom
        pythoncom.CoInitialize()
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(xls_path.absolute()))
        wb.SaveAs(str(xlsx_path.absolute()), FileFormat=51)  # 51 = xlsx
        wb.Close(False)
        excel.Quit()
        del wb
        del excel
        pythoncom.CoUninitialize()
        import gc
        gc.collect()
        import time
        time.sleep(0.5)
        logger.info("Convertido con Excel COM: %s", xls_path.name)
        return xlsx_path
    except ImportError:
        pass
    except Exception as e:
        logger.warning("Excel COM falló: %s", e)

    # Fallback: LibreOffice (si está instalado)
    for lo_path in [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        "soffice",
    ]:
        if shutil.which(lo_path) or Path(lo_path).exists():
            try:
                result = subprocess.run(
                    [lo_path, "--headless", "--convert-to", "xlsx",
                     str(xls_path), "--outdir", str(destino_dir)],
                    capture_output=True, text=True, timeout=60
                )
                if xlsx_path.exists():
                    logger.info("Convertido con LibreOffice: %s", xls_path.name)
                    return xlsx_path
            except Exception:
                pass

    logger.error("No se pudo convertir %s — instalar pywin32 o LibreOffice", xls_path.name)
    return None


# ==============================================================================
# EXTRACTOR
# ==============================================================================

def encontrar_hoja_alumnos(wb) -> str | None | list:
    """
    Encuentra la hoja de alumnos/matricula en el workbook.
    Retorna str para una sola hoja, list[str] para CB+CO (secundaria 2007-2015).
    """
    # 1. Candidatas exactas
    for candidata in HOJAS_CANDIDATAS:
        if candidata in wb.sheetnames:
            return candidata

    # 2. Estructura CB/CO (secundaria 2008-2015)
    hojas_cb = [h for h in wb.sheetnames if h.startswith("CB_") and "alu" in h.lower()]
    hojas_co = [h for h in wb.sheetnames if h.startswith("CO_") and "alu" in h.lower()]
    if hojas_cb and hojas_co:
        return [hojas_cb[0], hojas_co[0]]  # lista = sumar ambas

    # 3. Estructura Sec Bas/Orien (secundaria 2007)
    hojas_bas = [h for h in wb.sheetnames if "sec bas" in h.lower() and "alu" in h.lower()]
    hojas_orien = [h for h in wb.sheetnames if "sec orien" in h.lower() and "alu" in h.lower()]
    if hojas_bas and hojas_orien:
        return [hojas_bas[0], hojas_orien[0]]  # lista = sumar ambas

    # 3. Buscar hoja de matricula total (formato 2016+)
    EXCLUIR = re.compile(
        r"(promovid|repitient|egresad|pase|mujeres|secciones|cargos|horas|promoci|abandon|sobreedad)",
        re.IGNORECASE
    )
    candidatas_matricula = []
    for hoja in wb.sheetnames:
        nombre = hoja.lower()
        if ("matr" in nombre or "alu" in nombre) and not EXCLUIR.search(hoja):
            candidatas_matricula.append(hoja)

    if candidatas_matricula:
        return candidatas_matricula[0]

    return None


def extraer_datos(xlsx_path: Path, anio: int, nombre_nivel: str, codigo_nivel: str) -> pd.DataFrame:
    """Extrae datos de matrícula de un XLSX de anuario."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    nombre_hoja = encontrar_hoja_alumnos(wb)

    if not nombre_hoja:
        raise ValueError(f"No se encontró hoja de alumnos en {xlsx_path.name}. Hojas: {wb.sheetnames}")

    # Si es lista (CB+CO), procesar cada hoja y combinar resultados
    if isinstance(nombre_hoja, list):
        dfs = []
        for nh in nombre_hoja:
            ws = wb[nh]
            filas_tmp = list(ws.iter_rows(values_only=True))
            # Extraer marcadores y datos de esta sub-hoja
            marcadores_tmp = []
            for idx, fila in enumerate(filas_tmp):
                celda_a = str(fila[0]).strip() if fila[0] is not None else ""
                if PATRON_ESTATAL.search(celda_a):
                    marcadores_tmp.append(("Estatal", "E", idx))
                elif PATRON_PRIVADO.search(celda_a):
                    marcadores_tmp.append(("Privado", "P", idx))
            if len(marcadores_tmp) < 2:
                continue
            # Extraer datos de cada bloque
            for i, (sector_nombre, codigo_sector, idx_inicio) in enumerate(marcadores_tmp):
                idx_fin = marcadores_tmp[i+1][2] if i+1 < len(marcadores_tmp) else None
                bloque = filas_tmp[idx_inicio:idx_fin]
                idx_h = None
                for j, f in enumerate(bloque):
                    if f[0] and PATRON_HEADER.search(str(f[0])):
                        idx_h = j
                        break
                if idx_h is None:
                    continue
                for fila in bloque[idx_h+2:]:
                    jur = str(fila[0]).strip() if fila[0] else ""
                    if not jur or PATRON_RUIDO.match(jur):
                        continue
                    try:
                        total_raw = fila[1]
                        if total_raw is None or str(total_raw).strip() in ("", "-", "–", "—"):
                            continue
                        total = int(float(str(total_raw).replace(".", "").replace(",", ".")))
                        if total < 0:
                            continue
                    except (ValueError, TypeError):
                        continue
                    dfs.append({
                        "jurisdiccion_raw": jur,
                        "total_alumnos": total,
                        "sector": sector_nombre,
                        "codigo_sector": codigo_sector,
                        "anio_lectivo": anio,
                        "codigo_nivel_redfie": codigo_nivel,
                        "nivel_educativo": nombre_nivel,
                        "fuente_archivo": xlsx_path.name,
                        "version_etl": ETL_VERSION,
                    })
        if not dfs:
            return pd.DataFrame()
        # Sumar CB + CO por jurisdiccion y sector
        df_raw = pd.DataFrame(dfs)
        df_sum = df_raw.groupby(["jurisdiccion_raw","sector","codigo_sector","anio_lectivo",
                                  "codigo_nivel_redfie","nivel_educativo","fuente_archivo","version_etl"],
                                 as_index=False)["total_alumnos"].sum()
        return df_sum

    ws = wb[nombre_hoja]
    filas = list(ws.iter_rows(values_only=True))

    # Detectar marcadores de sector
    marcadores = []
    for idx, fila in enumerate(filas):
        celda_a = str(fila[0]).strip() if fila[0] is not None else ""
        if PATRON_ESTATAL.search(celda_a):
            marcadores.append(("Estatal", "E", idx))
        elif PATRON_PRIVADO.search(celda_a):
            marcadores.append(("Privado", "P", idx))

    if len(marcadores) < 2:
        logger.warning("Solo %d marcador(es) en %s — procesando como bloque único", len(marcadores), xlsx_path.name)

    if len(marcadores) < 2:
        logger.warning("Solo %d marcador(es) en %s — omitiendo archivo", len(marcadores), xlsx_path.name)
        return pd.DataFrame()

    resultados = []
    for i, (sector_nombre, codigo_sector, idx_inicio) in enumerate(marcadores):
        idx_fin = marcadores[i + 1][2] if i + 1 < len(marcadores) else None
        bloque = filas[idx_inicio:idx_fin]

        # Encontrar header
        idx_h = None
        for j, f in enumerate(bloque):
            if f[0] and PATRON_HEADER.search(str(f[0])):
                idx_h = j
                break

        if idx_h is None:
            logger.warning("Header no encontrado en bloque %s de %s", sector_nombre, xlsx_path.name)
            continue

        # Extraer datos: col 0 = jurisdiccion, col 1 = Total
        for fila in bloque[idx_h + 2:]:
            jur = str(fila[0]).strip() if fila[0] else ""
            if not jur or PATRON_RUIDO.match(jur):
                continue

            try:
                total_raw = fila[1]
                if total_raw is None or str(total_raw).strip() in ("", "-", "–", "—"):
                    continue
                total = int(float(str(total_raw).replace(".", "").replace(",", ".")))
                if total < 0:
                    continue
            except (ValueError, TypeError):
                continue

            resultados.append({
                "jurisdiccion_raw": jur,
                "total_alumnos": total,
                "sector": sector_nombre,
                "codigo_sector": codigo_sector,
                "anio_lectivo": anio,
                "codigo_nivel_redfie": codigo_nivel,
                "nivel_educativo": nombre_nivel,
                "fuente_archivo": xlsx_path.name,
                "version_etl": ETL_VERSION,
            })

    return pd.DataFrame(resultados)


# ==============================================================================
# CARGADOR SQL SERVER
# ==============================================================================

def get_conn():
    load_dotenv()
    server   = os.getenv("SQL_SERVER",   "localhost")
    database = os.getenv("SQL_DATABASE", "educacion_argentina")
    user     = os.getenv("SQL_USER",     "")
    password = os.getenv("SQL_PASSWORD", "")

    drivers = [d for d in pyodbc.drivers() if "SQL Server" in d]
    driver  = sorted(drivers, key=lambda d: int(re.search(r"\d+", d).group()) if re.search(r"\d+", d) else 0, reverse=True)[0]

    if user:
        cs = f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};UID={user};PWD={password};"
    else:
        cs = f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};Trusted_Connection=yes;"

    conn = pyodbc.connect(cs)
    conn.autocommit = False
    return conn


def cargar_dataframe(df: pd.DataFrame, conn) -> tuple[int, int]:
    """Carga el DataFrame en SQL Server. Retorna (insertadas, errores)."""
    cursor = conn.cursor()

    # Cache de dimensiones
    cursor.execute("SELECT codigo_redfie, nivel_id FROM dbo.dim_niveles_educativos")
    cache_nivel = {r[0]: r[1] for r in cursor.fetchall()}

    cursor.execute("SELECT codigo_sector, sector_id FROM dbo.dim_sectores")
    cache_sector = {r[0]: r[1] for r in cursor.fetchall()}

    cursor.execute("""
        SELECT LOWER(nombre_oficial), jurisdiccion_id FROM dbo.dim_jurisdicciones
        UNION
        SELECT LOWER(nombre_corto), jurisdiccion_id FROM dbo.dim_jurisdicciones
        WHERE nombre_corto IS NOT NULL
    """)
    cache_jur = {r[0]: r[1] for r in cursor.fetchall()}

    insertadas = 0
    errores = 0

    for _, row in df.iterrows():
        nivel_id  = cache_nivel.get(row["codigo_nivel_redfie"])
        sector_id = cache_sector.get(row["codigo_sector"])
        jur_id    = cache_jur.get(row["jurisdiccion_raw"].lower())

        if not nivel_id or not sector_id or not jur_id:
            if not jur_id:
                logger.debug("Jur no mapeada: '%s'", row["jurisdiccion_raw"])
            errores += 1
            continue

        try:
            cursor.execute(
                MERGE_SQL,
                int(row["anio_lectivo"]),
                jur_id, nivel_id, sector_id,
                int(row["total_alumnos"]),
                row["fuente_archivo"],
                row["version_etl"],
            )
            insertadas += 1
        except pyodbc.Error as e:
            logger.error("SQL error: %s", e)
            errores += 1

    return insertadas, errores


# ==============================================================================
# ORQUESTADOR PRINCIPAL
# ==============================================================================

def procesar_carpeta_anuario(carpeta: Path, anio: int, conn, base_dir: Path = None) -> dict:
    """Procesa todos los archivos de niveles en una carpeta de anuario."""
    resultado = {"anio": anio, "niveles": {}, "total_insertadas": 0, "total_errores": 0}

    # Buscar archivos relevantes (Inicial, Primario, Secundario, Superior)
    archivos = list(carpeta.rglob("*.xlsx")) + list(carpeta.rglob("*.xls")) + \
               list(carpeta.rglob("*.XLSX")) + list(carpeta.rglob("*.XLS"))

    archivos_niveles = []
    for a in archivos:
        nivel = inferir_nivel(a.stem)
        if nivel:
            archivos_niveles.append((a, nivel[0], nivel[1]))

    if not archivos_niveles:
        logger.warning("No se encontraron archivos de niveles en %s", carpeta)
        return resultado

    for archivo, nombre_nivel, codigo_nivel in archivos_niveles:
        logger.info("  Procesando: %s (%s)", archivo.name, nombre_nivel)

        # Convertir XLS si es necesario
        xlsx_path = archivo
        conv_dir = (base_dir or carpeta.parent) / "_xlsx_convertidos"
        conv_dir.mkdir(exist_ok=True)
        if archivo.suffix.lower() == ".xls":
            xlsx_conv = convertir_xls_a_xlsx(archivo, conv_dir)
            if not xlsx_conv:
                resultado["total_errores"] += 1
                continue
            xlsx_path = xlsx_conv

        try:
            df = extraer_datos(xlsx_path, anio, nombre_nivel, codigo_nivel)
            if df.empty:
                logger.warning("  DataFrame vacío: %s", archivo.name)
                continue

            insertadas, errores = cargar_dataframe(df, conn)
            resultado["niveles"][codigo_nivel] = {"insertadas": insertadas, "errores": errores}
            resultado["total_insertadas"] += insertadas
            resultado["total_errores"] += errores
            logger.info("  %s: %d insertadas, %d errores", nombre_nivel, insertadas, errores)

        except Exception as e:
            logger.error("  Error procesando %s: %s", archivo.name, e)
            resultado["total_errores"] += 1

        # No borrar el xlsx convertido — evita error de archivo bloqueado por Excel COM

    return resultado


def procesar_directorio_historico(directorio: str):
    """
    Punto de entrada principal.
    Procesa todos los anuarios en el directorio dado.
    """
    base = Path(directorio)
    if not base.exists():
        logger.error("El directorio '%s' no existe", directorio)
        sys.exit(1)

    logger.info("=" * 65)
    logger.info("ETL HISTÓRICO — %s", directorio)
    logger.info("=" * 65)

    conn = get_conn()

    resultados = []
    errores_globales = []

    # Procesar cada carpeta/archivo en el directorio
    for item in sorted(base.iterdir()):

        # Detectar año
        anio = inferir_anio(str(item))
        if not anio:
            logger.warning("No se pudo inferir año de '%s' — saltando", item.name)
            continue

        logger.info("\nProcesando año %d — %s", anio, item.name)

        # Si es archivo comprimido, descomprimir primero
        carpeta_trabajo = None
        if item.is_file() and item.suffix.lower() in (".zip", ".rar", ".7z"):
            carpeta_trabajo = base / f"_tmp_{item.stem}"
            carpeta_trabajo.mkdir(exist_ok=True)
            if not descomprimir(item, carpeta_trabajo):
                errores_globales.append(f"No se pudo descomprimir: {item.name}")
                continue
        elif item.is_dir():
            carpeta_trabajo = item
        else:
            continue

        try:
            resultado = procesar_carpeta_anuario(carpeta_trabajo, anio, conn, base_dir=base)
            conn.commit()
            resultados.append(resultado)
            logger.info("Año %d: %d filas cargadas", anio, resultado["total_insertadas"])
        except Exception as e:
            conn.rollback()
            logger.error("Error en año %d: %s", anio, e)
            errores_globales.append(f"Año {anio}: {e}")
        finally:
            # Limpiar carpeta temporal si se creó
            if carpeta_trabajo and carpeta_trabajo.name.startswith("_tmp_"):
                shutil.rmtree(carpeta_trabajo, ignore_errors=True)

    conn.close()

    # Resumen final
    logger.info("\n%s", "=" * 65)
    logger.info("RESUMEN FINAL")
    logger.info("=" * 65)
    total_global = sum(r["total_insertadas"] for r in resultados)
    logger.info("Años procesados: %d", len(resultados))
    logger.info("Total filas cargadas: %d", total_global)
    if errores_globales:
        logger.warning("Errores globales: %d", len(errores_globales))
        for e in errores_globales:
            logger.warning("  - %s", e)

    return resultados


# ==============================================================================
# ENTRY POINT
# ==============================================================================

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python etl_historico.py <carpeta_anuarios>")
        print('Ejemplo: python etl_historico.py "C:\\Users\\Administrator\\Desktop\\Anuarios"')
        sys.exit(1)

    load_dotenv()
    procesar_directorio_historico(sys.argv[1])
